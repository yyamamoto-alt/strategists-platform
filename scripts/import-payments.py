#!/usr/bin/env python3
"""
支払いDB.xlsx → Supabase (bank_transfers / payments テーブル) インポートスクリプト

既存テーブル bank_transfers, payments にデータをインポート/更新する。

使い方:
  1. pip install openpyxl supabase
  2. 環境変数を設定:
     export SUPABASE_URL="https://plrmqgcigzjuiovsbggf.supabase.co"
     export SUPABASE_SERVICE_KEY="eyJhbG..."
  3. python scripts/import-payments.py --dry-run  (確認)
  4. python scripts/import-payments.py             (本番投入)

オプション:
  --dry-run     SQLファイル出力のみ（DBには投入しない）
  --limit N     最初のN行だけ処理（テスト用）
  --clear       インポート前に既存データを削除（bank_transfers / payments 両方）
  --sheet bank  銀行シートのみ処理
  --sheet apps  Appsシートのみ処理
"""

import argparse
import os
import sys
import uuid
from datetime import datetime, date

import openpyxl

EXCEL_PATH = os.path.expanduser("~/Downloads/支払いDB.xlsx")

# ============================================================
# ユーティリティ（migrate-from-excel.py と同じパターン）
# ============================================================

def clean_value(val):
    """None, 空文字, '#N/A' 等を処理"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', 'None', '-'):
        return None
    return s


def to_date(val):
    """日付型への変換"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = clean_value(val)
    if s is None:
        return None
    try:
        return datetime.fromisoformat(s.replace(' ', 'T')).strftime('%Y-%m-%d')
    except Exception:
        pass
    return s


def to_int(val):
    """整数への変換"""
    s = clean_value(val)
    if s is None:
        return None
    try:
        f = float(s)
        return int(f)
    except Exception:
        return None


def to_text(val):
    """テキストへの変換"""
    s = clean_value(val)
    if s is None:
        return None
    if isinstance(val, (int, float)):
        if s.endswith('.0'):
            s = s[:-2]
    return s


def escape_sql(val):
    """SQL文字列エスケープ"""
    if val is None:
        return 'NULL'
    if isinstance(val, bool):
        return 'TRUE' if val else 'FALSE'
    if isinstance(val, (int, float)):
        return str(val)
    s = str(val).replace("'", "''")
    s = s.replace('\r\n', '\\n').replace('\r', '\\n').replace('\n', '\\n')
    s = s.replace('\t', '\\t')
    return f"E'{s}'"


# ============================================================
# 銀行シート処理
# ============================================================

def process_bank(wb, customer_email_map, limit=None):
    """銀行シートを処理して bank_transfers 用データリストを返す"""
    ws = wb['銀行']
    results = []
    skipped = 0

    # ヘッダー: 日付, 月, 購入者, 商品, 価格, 定価(税込), 割引後単価(税込), ジャンル, メアド, 状況
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # ヘッダー行スキップ
            continue
        if limit and i > limit:
            break

        row = list(row)

        # 購入者が空ならスキップ
        buyer_name = to_text(row[2] if len(row) > 2 else None)
        if not buyer_name:
            skipped += 1
            continue

        email = to_text(row[8] if len(row) > 8 else None)
        customer_id = None
        if email and customer_email_map:
            customer_id = customer_email_map.get(email.lower())

        data = {
            'id': str(uuid.uuid4()),
            'transfer_date': to_date(row[0] if len(row) > 0 else None),
            'period': to_date(row[1] if len(row) > 1 else None),
            'buyer_name': buyer_name,
            'product': to_text(row[3] if len(row) > 3 else None),
            'amount': to_int(row[4] if len(row) > 4 else None),
            'list_price': to_int(row[5] if len(row) > 5 else None),
            'discounted_price': to_int(row[6] if len(row) > 6 else None),
            'genre': to_text(row[7] if len(row) > 7 else None),
            'email': email,
            'status': to_text(row[9] if len(row) > 9 else None),
            'customer_id': customer_id,
        }
        results.append(data)

    print(f"  銀行: {len(results)} records, {skipped} skipped")
    return results


# ============================================================
# Appsシート処理
# ============================================================

def process_apps(wb, customer_email_map, limit=None):
    """Appsシートを処理して payments 用データリストを返す"""
    ws = wb['Apps']
    results = []
    skipped = 0

    # ヘッダー: プラン名, 種別, メールアドレス, 名前, 購入日・契約日, 状態,
    #           販売・契約金額, 次回決済予定日, 備考, 分割価格, 分割回数, 年月
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # ヘッダー行スキップ
            continue
        if limit and i > limit:
            break

        row = list(row)

        # プラン名が空ならスキップ
        plan_name = to_text(row[0] if len(row) > 0 else None)
        if not plan_name:
            skipped += 1
            continue

        email = to_text(row[2] if len(row) > 2 else None)
        customer_id = None
        if email and customer_email_map:
            customer_id = customer_email_map.get(email.lower())

        data = {
            'id': str(uuid.uuid4()),
            'plan_name': plan_name,
            'payment_type': to_text(row[1] if len(row) > 1 else None),
            'email': email,
            'customer_name': to_text(row[3] if len(row) > 3 else None),
            'purchase_date': to_date(row[4] if len(row) > 4 else None),
            'status': to_text(row[5] if len(row) > 5 else None),
            'amount': to_int(row[6] if len(row) > 6 else None),
            'next_billing_date': to_text(row[7] if len(row) > 7 else None),
            'memo': to_text(row[8] if len(row) > 8 else None),
            'installment_amount': to_int(row[9] if len(row) > 9 else None),
            'installment_count': to_int(row[10] if len(row) > 10 else None),
            'period': to_text(row[11] if len(row) > 11 else None),
            'customer_id': customer_id,
        }
        results.append(data)

    print(f"  Apps: {len(results)} records, {skipped} skipped")
    return results


# ============================================================
# 顧客マスタ取得
# ============================================================

def fetch_customer_email_map(supabase_url, supabase_key):
    """Supabaseからcustomers.email → id のマッピングを取得"""
    try:
        from supabase import create_client
        client = create_client(supabase_url, supabase_key)

        result = client.table('customers').select('id, email').execute()
        email_map = {}
        for row in result.data:
            if row.get('email'):
                email_map[row['email'].lower()] = row['id']
        print(f"  顧客マスタ: {len(email_map)} emails loaded")
        return email_map
    except Exception as e:
        print(f"  警告: 顧客マスタ取得失敗 ({e}), customer_id は設定されません")
        return {}


# ============================================================
# SQL出力
# ============================================================

def dict_to_insert_sql(table, data):
    """dictからINSERT文を生成"""
    cols = []
    vals = []
    for k, v in data.items():
        if v is not None:
            cols.append(k)
            vals.append(escape_sql(v))
    return f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(vals)});"


def generate_sql(table_name, records, output_dir):
    """SQL出力"""
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f'{table_name}.sql')

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f"-- {table_name} ({len(records)} records)\n")
        f.write(f"-- 生成日時: {datetime.now().isoformat()}\n\n")
        f.write("BEGIN;\n\n")
        for data in records:
            f.write(dict_to_insert_sql(table_name, data) + '\n')
        f.write("\nCOMMIT;\n")

    print(f"  SQL出力: {filepath} ({len(records)} records)")
    return filepath


# ============================================================
# Supabase投入
# ============================================================

def import_to_supabase(table_name, records, supabase_url, supabase_key, clear=False):
    """Supabaseに直接投入"""
    from supabase import create_client
    client = create_client(supabase_url, supabase_key)

    if clear:
        print(f"  {table_name}: 既存データ削除中...")
        client.table(table_name).delete().neq('id', '00000000-0000-0000-0000-000000000000').execute()
        print(f"  {table_name}: 削除完了")

    # 500件ずつバッチ投入
    batch_size = 500
    total = len(records)
    inserted = 0

    for i in range(0, total, batch_size):
        batch = records[i:i + batch_size]
        # None値のキーを除去
        clean_batch = []
        for record in batch:
            clean_record = {k: v for k, v in record.items() if v is not None}
            clean_batch.append(clean_record)

        try:
            client.table(table_name).insert(clean_batch).execute()
            inserted += len(clean_batch)
            print(f"  {table_name}: 投入 {inserted}/{total} ({inserted * 100 // total}%)")
        except Exception as e:
            print(f"  エラー (行 {i+1}-{i+len(batch)}): {e}")
            # 個別に投入を試みる
            for j, record in enumerate(clean_batch):
                try:
                    client.table(table_name).insert(record).execute()
                    inserted += 1
                except Exception as e2:
                    print(f"    スキップ (行 {i+j+1}): {e2}")

    print(f"  {table_name}: 完了 {inserted}/{total} records inserted")
    return inserted


def main():
    parser = argparse.ArgumentParser(description='支払いDB → Supabase インポート')
    parser.add_argument('--excel', default=EXCEL_PATH, help='Excelファイルパス')
    parser.add_argument('--dry-run', action='store_true', help='SQLファイル出力のみ')
    parser.add_argument('--limit', type=int, help='処理行数制限（テスト用）')
    parser.add_argument('--clear', action='store_true', help='インポート前に既存データ削除')
    parser.add_argument('--sheet', choices=['bank', 'apps'], help='特定シートのみ処理')
    args = parser.parse_args()

    supabase_url = os.environ.get('SUPABASE_URL', '')
    supabase_key = os.environ.get('SUPABASE_SERVICE_KEY', '')

    print(f"Excelファイル: {args.excel}")
    print(f"モード: {'dry-run (SQL出力のみ)' if args.dry_run else '本番投入'}")
    if args.limit:
        print(f"制限: 最初の{args.limit}行のみ")
    if args.sheet:
        print(f"対象シート: {args.sheet}")
    print()

    # 顧客メールマップ取得
    customer_email_map = {}
    if supabase_url and supabase_key:
        print("顧客マスタ読み込み中...")
        customer_email_map = fetch_customer_email_map(supabase_url, supabase_key)
    else:
        print("警告: SUPABASE_URL/SUPABASE_SERVICE_KEY未設定 → customer_id は設定されません")

    # Excel読み込み
    print(f"\nExcelファイル読み込み中...")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    bank_records = []
    apps_records = []

    if not args.sheet or args.sheet == 'bank':
        print(f"\n銀行シート処理中...")
        bank_records = process_bank(wb, customer_email_map, limit=args.limit)

    if not args.sheet or args.sheet == 'apps':
        print(f"\nAppsシート処理中...")
        apps_records = process_apps(wb, customer_email_map, limit=args.limit)

    wb.close()

    if args.dry_run:
        output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'migration_sql')
        if bank_records:
            generate_sql('bank_transfers_new', bank_records, output_dir)
        if apps_records:
            generate_sql('payments_new', apps_records, output_dir)
    else:
        if not supabase_url or not supabase_key:
            print("\nエラー: SUPABASE_URL と SUPABASE_SERVICE_KEY を設定してください")
            sys.exit(1)

        total_inserted = 0
        if bank_records:
            total_inserted += import_to_supabase(
                'bank_transfers', bank_records, supabase_url, supabase_key, clear=args.clear
            )
        if apps_records:
            total_inserted += import_to_supabase(
                'payments', apps_records, supabase_url, supabase_key, clear=args.clear
            )

        print(f"\n合計 {total_inserted} レコード投入完了")

    total = len(bank_records) + len(apps_records)
    print(f"\n合計 {total} レコード処理完了 (銀行: {len(bank_records)}, Apps: {len(apps_records)})")


if __name__ == '__main__':
    main()
