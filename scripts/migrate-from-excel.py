#!/usr/bin/env python3
"""
スプレッドシート(Excel) → Supabase 移行スクリプト

使い方:
  1. pip install openpyxl supabase
  2. 環境変数を設定:
     export SUPABASE_URL="https://xxxxx.supabase.co"
     export SUPABASE_SERVICE_KEY="eyJhbG..."
  3. python scripts/migrate-from-excel.py --excel "path/to/経営管理.xlsx" --dry-run
  4. 確認後: python scripts/migrate-from-excel.py --excel "path/to/経営管理.xlsx"

オプション:
  --dry-run    SQLをファイルに出力するだけ（DBには投入しない）
  --limit N    最初のN行だけ処理（テスト用）
"""

import argparse
import json
import os
import sys
import uuid
from datetime import datetime, date
from decimal import Decimal

import openpyxl

# ============================================================
# カラムマッピング定義
# Excel列番号(1-indexed) → Supabaseテーブル.カラム名
# ============================================================

CUSTOMER_MAPPING = {
    1: 'application_date',       # 申込日
    2: 'name',                   # 名前
    3: 'email',                  # メール
    4: 'phone',                  # 電話番号
    5: 'utm_source',             # utm_source
    6: 'utm_medium',             # utm_medium
    7: 'utm_id',                 # utm_id
    8: 'utm_campaign',           # utm_campaign
    9: 'attribute',              # 属性
    10: 'career_history',        # 経歴
    17: 'initial_channel',       # 初回認知経路 → reference_media にも入れるが分離
    84: 'karte_email',           # メアド（カルテ）
    85: 'karte_phone',           # 電話番号(カルテ)
    86: 'birth_date',            # 生年月日
    87: 'name_kana',             # フリガナ
    88: 'target_companies',      # 志望企業
    89: 'target_firm_type',      # 対策ファームのご意向
    90: 'initial_level',         # 申込時レベル
    93: 'application_reason_karte',  # 申込の決め手（カルテ）
    94: 'program_interest',      # 有料プログラムへの関心
    95: 'desired_schedule',      # 希望期間・頻度
    96: 'purchased_content',     # ご購入いただいたコンテンツ
    97: 'parent_support',        # 親御様からの支援
    98: 'sns_accounts',          # 就活アカウント(X)
    99: 'reference_media',       # 参考メディア
    100: 'hobbies',              # 趣味・特技
    101: 'behavioral_traits',    # 行動特性
    102: 'other_background',     # その他要望・特記事項
    103: 'notes',                # 備考
    104: 'caution_notes',        # 注意事項
    130: 'transfer_intent',      # 転職意向
    131: 'university',           # 大学名抜き出し
}

PIPELINE_MAPPING = {
    11: 'agent_interest_at_application',  # 申込時点エージェント (text→bool later)
    12: 'meeting_scheduled_date',  # 面接予定時期 (テキストの場合あり)
    13: 'stage',                 # 検討状況 → stage
    14: 'projected_amount',      # 売上(見込)
    15: 'decision_factor',       # 検討・失注理由
    16: 'deal_status',           # 実施状況
    18: 'sales_content',         # 申し込みの決め手 → decision_factor的
    19: 'sales_date',            # 営業実施日
    20: 'probability',           # 確度
    21: 'response_date',         # 返答日/仮入会日
    22: 'sales_person',          # 営業担当
    23: 'sales_content',         # 営業内容 (上書き注意→Col18とマージ)
    24: 'sales_strategy',        # 営業方針
    25: 'jicoo_message',         # jicooメッセージ
    26: 'agent_confirmation',    # エージェント利用意向
    27: 'marketing_memo',        # マーケメモ
    28: 'sales_route',           # 経路(営業担当記入)
    29: 'comparison_services',   # 比較サービス
    30: 'first_reward_category', # 一次報酬分類
    31: 'performance_reward_category',  # 成果報酬分類
    32: 'lead_time',             # 入会までのリードタイム
    33: 'google_ads_target',     # Google広告成果対象
    17: 'initial_channel',       # 初回認知経路
    122: 'alternative_application',  # 別経由での応募の有無
    133: 'additional_sales_content',  # [追加指導] 営業内容
    134: 'additional_plan',      # [追加指導]プラン
    135: 'additional_discount_info',  # [追加指導]割引制度の案内
    136: 'additional_notes',     # [追加指導]学び
}

CONTRACT_MAPPING = {
    34: 'referral_category',     # 人材紹介区分
    35: 'referral_status',       # 紹介ステータス
    36: 'first_amount',          # 一次報酬請求予定額
    37: 'confirmed_amount',      # 確定売上
    38: 'discount',              # 割引
    39: 'progress_sheet_url',    # Progress Sheet
    40: 'enrollment_status',     # 受講状況
    41: 'plan_name',             # 受講サービス名
    43: 'payment_date',          # 入金日
    119: 'invoice_info',         # 請求書用
    121: 'billing_status',       # 請求状況
    139: 'subsidy_eligible',     # リスキャリ補助金対象 (text→bool)
    140: 'subsidy_amount',       # 補助金額
}

LEARNING_MAPPING = {
    42: 'mentor_name',           # 指導メンター
    44: 'coaching_start_date',   # 指導開始日
    45: 'coaching_end_date',     # 指導終了日
    46: 'last_coaching_date',    # 最終指導日
    47: 'contract_months',       # 契約月数
    48: 'total_sessions',        # 契約指導回数
    49: 'weekly_sessions',       # 週あたり指導数
    50: 'completed_sessions',    # 指導完了数
    51: 'attendance_rate',       # 日程消化率
    52: 'session_completion_rate',  # 指導消化率
    53: 'progress_text',         # 進捗
    54: 'level_fermi',           # 最新レベル(フェルミ)
    55: 'level_case',            # 最新レベル(ケース)
    56: 'level_mck',             # 最新レベル(McK)
    58: 'selection_status',      # 選考状況
    59: 'level_up_range',        # レベルアップ幅
    60: 'interview_timing_at_end',  # 面接予定時期（指導終了時点）
    61: 'target_companies_at_end',  # 受験企業（指導終了時点）
    62: 'offer_probability_at_end',  # 内定確度判定
    63: 'additional_coaching_proposal',  # 追加指導提案
    64: 'initial_coaching_level',  # 指導開始時レベル
    65: 'enrollment_form_date',  # 入会フォーム提出日
    66: 'coaching_requests',     # 指導要望
    67: 'enrollment_reason',     # 入会理由
    69: 'behavior_session1',     # ビヘイビア1回目
    70: 'behavior_session2',     # ビヘイビア2回目
    71: 'assessment_session1',   # アセスメント1回目
    72: 'assessment_session2',   # アセスメント2回目
    74: 'extension_days',        # 延長分(日)
    91: 'case_interview_progress',  # ケース面接対策の進捗
    92: 'case_interview_weaknesses',  # ケース面接で苦手なこと
    111: 'mentoring_satisfaction',  # メンタリング満足度
    132: 'start_email_sent',     # 受講開始日メール送付済み
}

AGENT_MAPPING = {
    68: 'agent_memo',            # エージェント業務メモ
    73: 'expected_agent_revenue',  # 人材見込売上
    75: 'offer_company',         # 就活/転職活動の結果・内定先
    76: 'external_agents',       # 利用エージェント
    77: 'hire_rate',             # 入社至る率
    78: 'offer_probability',     # 内定確度
    79: 'offer_salary',          # 想定年収
    80: 'referral_fee_rate',     # 紹介料率
    81: 'margin',                # マージン
    82: 'placement_date',        # 入社予定日
    83: 'general_memo',          # メモ
    118: 'loss_reason',          # エージェント失注理由
    128: 'expected_referral_fee',  # 人材紹介報酬期待値
    137: 'agent_staff',          # エージェント担当者
    141: 'placement_confirmed',  # 人材確定
}

# ============================================================
# ユーティリティ
# ============================================================

def clean_value(val):
    """None, 空文字, '#N/A' 等を処理"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('', '#N/A', '#REF!', '#VALUE!', '#DIV/0!', '#NAME?', 'None', '-'):
        return None
    return s


def to_date(val):
    """日付型への変換"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = clean_value(val)
    if s is None:
        return None
    # "2026-02-18 00:00:00" 形式
    try:
        return datetime.fromisoformat(s.replace(' ', 'T')).strftime('%Y-%m-%d')
    except:
        pass
    return s  # テキストとして保存 (例: "5月")


def to_int(val):
    """整数への変換"""
    s = clean_value(val)
    if s is None:
        return None
    try:
        f = float(s)
        return int(f)
    except:
        return None


def to_float(val):
    """小数への変換"""
    s = clean_value(val)
    if s is None:
        return None
    try:
        return float(s)
    except:
        return None


def to_text(val):
    """テキストへの変換"""
    s = clean_value(val)
    if s is None:
        return None
    # 数値が電話番号として入っている場合
    if isinstance(val, (int, float)):
        if s.endswith('.0'):
            s = s[:-2]
    return s


def to_bool_text(val):
    """'対象' → True, etc."""
    s = clean_value(val)
    if s is None:
        return None
    if s in ('対象', 'TRUE', 'True', 'true', '1', 'はい', 'Yes'):
        return True
    return False


def escape_sql(val):
    """SQL文字列エスケープ"""
    if val is None:
        return 'NULL'
    if isinstance(val, bool):
        return 'TRUE' if val else 'FALSE'
    if isinstance(val, (int, float)):
        return str(val)
    s = str(val).replace("'", "''")
    # 改行・タブ等をエスケープ（SQLが途切れないように）
    s = s.replace('\r\n', '\\n').replace('\r', '\\n').replace('\n', '\\n')
    s = s.replace('\t', '\\t')
    return f"E'{s}'"


# ============================================================
# 行データから各テーブル用のdictを作る
# ============================================================

def extract_customer(row, customer_id):
    """顧客DB行 → customers テーブル用dict"""
    data = {'id': customer_id}
    for col_idx, db_col in CUSTOMER_MAPPING.items():
        val = row[col_idx - 1] if col_idx <= len(row) else None

        if db_col in ('application_date', 'birth_date'):
            t = to_text(val)
            if t:
                d = to_date(val)
                data[db_col] = d if d else t
            else:
                data[db_col] = None
        elif db_col == 'phone':
            data[db_col] = to_text(val)
        elif db_col == 'attribute':
            data[db_col] = to_text(val)
        else:
            data[db_col] = to_text(val)

    return data


def extract_pipeline(row, customer_id):
    """顧客DB行 → sales_pipeline テーブル用dict"""
    data = {'customer_id': customer_id}

    for col_idx, db_col in PIPELINE_MAPPING.items():
        val = row[col_idx - 1] if col_idx <= len(row) else None

        if db_col == 'agent_interest_at_application':
            t = to_text(val)
            if t:
                data[db_col] = t  # テキストとして保存
        elif db_col in ('sales_date', 'response_date', 'status_confirmed_date', 'status_final_date'):
            # 日付変換を試み、失敗したらテキストのまま保存
            t = to_text(val)
            if t:
                d = to_date(val)
                data[db_col] = d if d else t
            else:
                data[db_col] = None
        elif db_col in ('projected_amount',):
            data[db_col] = to_int(val)
        elif db_col == 'probability':
            data[db_col] = to_float(val)
        elif db_col == 'stage':
            t = to_text(val)
            data[db_col] = t or '問い合わせ'
        elif db_col == 'deal_status':
            t = to_text(val)
            data[db_col] = t or '未対応'
        else:
            data[db_col] = to_text(val)

    # Col18 と Col23 が両方 sales_content にマップされる問題を解決
    col18 = to_text(row[17] if len(row) > 17 else None)
    col23 = to_text(row[22] if len(row) > 22 else None)
    if col18 and col23:
        data['decision_factor'] = col18
        data['sales_content'] = col23
    elif col18:
        data['decision_factor'] = col18
    elif col23:
        data['sales_content'] = col23

    return data


def extract_contract(row, customer_id):
    """顧客DB行 → contracts テーブル用dict"""
    data = {'customer_id': customer_id}

    for col_idx, db_col in CONTRACT_MAPPING.items():
        val = row[col_idx - 1] if col_idx <= len(row) else None

        if db_col == 'payment_date':
            t = to_text(val)
            if t:
                d = to_date(val)
                data[db_col] = d if d else t
            else:
                data[db_col] = None
        elif db_col in ('first_amount', 'confirmed_amount', 'subsidy_amount'):
            data[db_col] = to_int(val)
        elif db_col == 'discount':
            t = to_text(val)
            # 割引は金額の場合とテキストの場合がある
            i = to_int(val)
            data['discount'] = i  # 数値なら数値
            if i is None and t:
                # テキスト（"案内なし"等）の場合は0にしてメモに保存
                data['discount'] = 0
        elif db_col == 'subsidy_eligible':
            data[db_col] = to_bool_text(val)
        elif db_col == 'billing_status':
            t = to_text(val)
            data[db_col] = t or '未請求'
        else:
            data[db_col] = to_text(val)

    return data


def extract_learning(row, customer_id):
    """顧客DB行 → learning_records テーブル用dict"""
    data = {'customer_id': customer_id}

    for col_idx, db_col in LEARNING_MAPPING.items():
        val = row[col_idx - 1] if col_idx <= len(row) else None

        if db_col in ('coaching_start_date', 'coaching_end_date', 'last_coaching_date', 'enrollment_form_date'):
            t = to_text(val)
            if t:
                d = to_date(val)
                data[db_col] = d if d else t
            else:
                data[db_col] = None
        elif db_col in ('total_sessions', 'completed_sessions', 'contract_months', 'extension_days'):
            data[db_col] = to_int(val)
        elif db_col in ('attendance_rate', 'session_completion_rate', 'weekly_sessions'):
            data[db_col] = to_float(val)
        else:
            data[db_col] = to_text(val)

    return data


def extract_agent(row, customer_id):
    """顧客DB行 → agent_records テーブル用dict"""
    data = {'customer_id': customer_id}

    for col_idx, db_col in AGENT_MAPPING.items():
        val = row[col_idx - 1] if col_idx <= len(row) else None

        if db_col == 'placement_date':
            t = to_text(val)
            if t:
                d = to_date(val)
                data[db_col] = d if d else t
            else:
                data[db_col] = None
        elif db_col in ('expected_agent_revenue', 'offer_salary', 'margin', 'expected_referral_fee'):
            data[db_col] = to_int(val)
        elif db_col in ('hire_rate', 'offer_probability', 'referral_fee_rate'):
            data[db_col] = to_float(val)
        else:
            data[db_col] = to_text(val)

    return data


# ============================================================
# SQL生成
# ============================================================

def dict_to_insert_sql(table, data):
    """dictからINSERT文を生成"""
    cols = []
    vals = []
    for k, v in data.items():
        if v is not None:
            cols.append(k)
            vals.append(escape_sql(v))
    return f"INSERT INTO {table} ({', '.join(cols)}) VALUES ({', '.join(vals)});"


# ============================================================
# メイン処理
# ============================================================

def process_customer_db(wb, limit=None):
    """顧客DBシートを処理"""
    ws = wb['顧客DB(new)']
    results = {
        'customers': [],
        'sales_pipeline': [],
        'contracts': [],
        'learning_records': [],
        'agent_records': [],
    }
    customer_email_to_id = {}
    skipped = 0

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # ヘッダー行スキップ
            continue
        if limit and i > limit:
            break

        row = list(row)

        # 名前が空ならスキップ
        name = to_text(row[1] if len(row) > 1 else None)
        if not name:
            skipped += 1
            continue

        customer_id = str(uuid.uuid4())
        email = to_text(row[2] if len(row) > 2 else None)
        if email:
            customer_email_to_id[email.lower()] = customer_id

        # 各テーブル用データを抽出
        customer_data = extract_customer(row, customer_id)
        results['customers'].append(customer_data)

        pipeline_data = extract_pipeline(row, customer_id)
        if pipeline_data:
            pipeline_data['id'] = str(uuid.uuid4())
            results['sales_pipeline'].append(pipeline_data)

        contract_data = extract_contract(row, customer_id)
        if contract_data:
            contract_data['id'] = str(uuid.uuid4())
            results['contracts'].append(contract_data)

        learning_data = extract_learning(row, customer_id)
        if learning_data:
            learning_data['id'] = str(uuid.uuid4())
            results['learning_records'].append(learning_data)

        agent_data = extract_agent(row, customer_id)
        if agent_data:
            agent_data['id'] = str(uuid.uuid4())
            results['agent_records'].append(agent_data)

    print(f"  顧客DB: {len(results['customers'])} customers, {skipped} skipped")
    print(f"  Pipeline: {len(results['sales_pipeline'])} records")
    print(f"  Contracts: {len(results['contracts'])} records")
    print(f"  Learning: {len(results['learning_records'])} records")
    print(f"  Agent: {len(results['agent_records'])} records")

    return results, customer_email_to_id


def process_apps(wb, customer_email_to_id, limit=None):
    """Appsシートを処理"""
    ws = wb['Apps']
    results = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if limit and i > limit:
            break

        row = list(row)
        email = to_text(row[2] if len(row) > 2 else None)
        customer_id = customer_email_to_id.get(email.lower()) if email else None

        data = {
            'id': str(uuid.uuid4()),
            'plan_name': to_text(row[0] if len(row) > 0 else None),
            'payment_type': to_text(row[1] if len(row) > 1 else None),
            'email': email,
            'customer_name': to_text(row[3] if len(row) > 3 else None),
            'purchase_date': to_date(row[4] if len(row) > 4 else None),
            'status': to_text(row[5] if len(row) > 5 else None),
            'amount': to_int(row[6] if len(row) > 6 else None),
            'next_billing_date': to_text(row[7] if len(row) > 7 else None),
            'memo': to_text(row[8] if len(row) > 8 else None),
            'installment_amount': to_int(row[9] if len(row) > 9 else None),
            'installment_count': to_int(row[10] if len(row) > 10 else None),
            'period': to_text(row[11] if len(row) > 11 else None),
            'customer_id': customer_id,
        }
        results.append(data)

    print(f"  Apps: {len(results)} payment records")
    return results


def process_bank(wb, customer_email_to_id, limit=None):
    """銀行シートを処理"""
    ws = wb['銀行']
    results = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if limit and i > limit:
            break

        row = list(row)
        email = to_text(row[8] if len(row) > 8 else None)
        customer_id = customer_email_to_id.get(email.lower()) if email else None

        data = {
            'id': str(uuid.uuid4()),
            'transfer_date': to_date(row[0] if len(row) > 0 else None),
            'period': to_date(row[1] if len(row) > 1 else None),
            'buyer_name': to_text(row[2] if len(row) > 2 else None),
            'product': to_text(row[3] if len(row) > 3 else None),
            'amount': to_int(row[4] if len(row) > 4 else None),
            'list_price': to_int(row[5] if len(row) > 5 else None),
            'discounted_price': to_int(row[6] if len(row) > 6 else None),
            'genre': to_text(row[7] if len(row) > 7 else None),
            'email': email,
            'status': to_text(row[9] if len(row) > 9 else None),
            'customer_id': customer_id,
        }
        results.append(data)

    print(f"  銀行: {len(results)} bank transfer records")
    return results


CHUNK_SIZE = 500  # 1ファイルあたりの最大レコード数


def write_table_sql(dir_path, seq, table_name, records):
    """テーブル別にSQL分割ファイルを出力（500件ずつチャンク分割）"""
    if len(records) <= CHUNK_SIZE:
        filename = f"{seq:02d}_{table_name}.sql"
        filepath = os.path.join(dir_path, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"-- {table_name} ({len(records)} records)\n")
            f.write(f"-- 生成日時: {datetime.now().isoformat()}\n\n")
            f.write("BEGIN;\n\n")
            for data in records:
                f.write(dict_to_insert_sql(table_name, data) + '\n')
            f.write("\nCOMMIT;\n")
        print(f"  {filename}: {len(records)} records")
    else:
        chunks = [records[i:i + CHUNK_SIZE] for i in range(0, len(records), CHUNK_SIZE)]
        for ci, chunk in enumerate(chunks):
            suffix = chr(ord('a') + ci)  # a, b, c, ...
            filename = f"{seq:02d}{suffix}_{table_name}.sql"
            filepath = os.path.join(dir_path, filename)
            start = ci * CHUNK_SIZE + 1
            end = start + len(chunk) - 1
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(f"-- {table_name} (records {start}-{end} / {len(records)})\n")
                f.write(f"-- 生成日時: {datetime.now().isoformat()}\n\n")
                f.write("BEGIN;\n\n")
                for data in chunk:
                    f.write(dict_to_insert_sql(table_name, data) + '\n')
                f.write("\nCOMMIT;\n")
            print(f"  {filename}: {len(chunk)} records ({start}-{end})")


def generate_sql(results, payments, bank_transfers, output_path):
    """テーブル別にSQL分割ファイルを出力"""
    dir_path = os.path.dirname(output_path) or 'scripts'
    sql_dir = os.path.join(dir_path, 'migration_sql')
    os.makedirs(sql_dir, exist_ok=True)

    print(f"\nSQL分割出力先: {sql_dir}/")

    # 0: クリーンアップSQL
    cleanup_path = os.path.join(sql_dir, '00_cleanup.sql')
    with open(cleanup_path, 'w', encoding='utf-8') as f:
        f.write("-- 既存データ削除（外部キー制約の順序で）\n")
        f.write(f"-- 生成日時: {datetime.now().isoformat()}\n\n")
        f.write("DELETE FROM agent_records;\n")
        f.write("DELETE FROM learning_records;\n")
        f.write("DELETE FROM contracts;\n")
        f.write("DELETE FROM sales_pipeline;\n")
        f.write("DELETE FROM payments;\n")
        f.write("DELETE FROM bank_transfers;\n")
        f.write("DELETE FROM customers;\n")
    print(f"  00_cleanup.sql: DELETE文")

    # テーブル別に分割出力
    tables = [
        (1, 'customers', results['customers']),
        (2, 'sales_pipeline', results['sales_pipeline']),
        (3, 'contracts', results['contracts']),
        (4, 'learning_records', results['learning_records']),
        (5, 'agent_records', results['agent_records']),
        (6, 'payments', payments),
        (7, 'bank_transfers', bank_transfers),
    ]

    for seq, table_name, records in tables:
        write_table_sql(sql_dir, seq, table_name, records)

    print(f"\n実行順序:")
    print(f"  1. 00_cleanup.sql      ← 既存データ削除")
    print(f"  2. 01_customers.sql    ← 顧客（先に投入: 外部キー参照元）")
    print(f"  3. 02〜07 を順番に投入")


def main():
    parser = argparse.ArgumentParser(description='スプレッドシート → Supabase 移行')
    parser.add_argument('--excel', required=True, help='Excelファイルパス')
    parser.add_argument('--dry-run', action='store_true', help='SQLファイル出力のみ')
    parser.add_argument('--limit', type=int, help='処理行数制限（テスト用）')
    parser.add_argument('--output', default='scripts/migration_data.sql', help='SQL出力先')
    args = parser.parse_args()

    print(f"📂 Excelファイル: {args.excel}")
    print(f"🔧 モード: {'dry-run (SQL出力のみ)' if args.dry_run else '本番投入'}")
    if args.limit:
        print(f"⚠️  制限: 最初の{args.limit}行のみ")

    print("\n📖 Excelファイル読み込み中...")
    wb = openpyxl.load_workbook(args.excel, read_only=True, data_only=True)

    print("\n📊 顧客DB処理中...")
    results, email_map = process_customer_db(wb, limit=args.limit)

    print("\n💳 Apps処理中...")
    payments = process_apps(wb, email_map, limit=args.limit)

    print("\n🏦 銀行処理中...")
    bank_transfers = process_bank(wb, email_map, limit=args.limit)

    wb.close()

    # SQL出力
    generate_sql(results, payments, bank_transfers, args.output)

    total = (
        len(results['customers']) +
        len(results['sales_pipeline']) +
        len(results['contracts']) +
        len(results['learning_records']) +
        len(results['agent_records']) +
        len(payments) +
        len(bank_transfers)
    )
    print(f"\n✅ 合計 {total} レコード")

    if not args.dry_run:
        print("\n⚠️  Supabase投入は別途実装予定")
        print("現在はSQL出力のみ対応しています")
        print(f"生成されたSQLを確認して、Supabase SQL Editorで実行してください:")
        print(f"  {args.output}")


if __name__ == '__main__':
    main()
